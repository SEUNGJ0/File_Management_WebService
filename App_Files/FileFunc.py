from openpyxl import load_workbook, styles
import os
import shutil
import datetime
import zipfile

class File_Manager():
    def Absorption(file_list, name = None):
        # 추출할 파일들의 경로 설정
        path = os.getcwd()+"/media/file/관리/취합 파일/"
        colonm = ['A','B','C','D','E','F','G','H','I']
        date = datetime.datetime.now().strftime("%m월%d일%H시%M분")
        results = []
        error_messages = {}
        
        # 파일 리스트의 각 파일에서 데이터를 추출해서 리스트로 통합 저장
        for file_name_raw in file_list:
            error_message = []
            file_name = path + file_name_raw
            wb = load_workbook(filename=file_name, data_only=True)
            ws = wb.active
            for i in range(3,ws.max_row): # 읽을 행 지정
                result = []
                if ws['A'+str(i)].value:
                    for j in colonm: # 열 지정
                        if j in ['A','B','C','G','H','I']:
                            if type(ws[j+str(i)].value) == int:
                                result.append(ws[j+str(i)].value)
                            else:
                                error_message.append(f"{j}{i}")
                        else:
                            if type(ws[j+str(i)].value) == str:
                                result.append(ws[j+str(i)].value)
                            else:
                                error_message.append(f"{j}{i}")
                else:
                    break
                results.append(result)
            if error_message:
                error_messages[file_name_raw] = error_message 
        
        if not error_messages:
            guide = os.getcwd()+"/media/file/공지사항/admin/자료취합_양식_파일.xlsx"
            wb = load_workbook(filename=guide)
            ws = wb.active
            count_R = 3
            for line in results:
                count_C = 0
                for cell in line:
                    ws[colonm[count_C]+str(count_R)].value = cell
                    ws[colonm[count_C]+str(count_R)].alignment = styles.Alignment(horizontal = 'center', vertical='center') # 셀의 텍스트 위치 조정
                    count_C += 1
                count_R += 1

            if name :
                save_path = os.getcwd()+"/media/file/관리/통합 파일/"+str(name)+".xlsx"
            else:
                save_path = os.getcwd()+"/media/file/관리/통합 파일/"+"통합 파일_"+date+".xlsx"
                path = "file/관리/통합 파일/"+"통합 파일_"+date+".xlsx"
            wb.save(save_path)
            return path
        else :
            return error_messages

    def CopyAndMove(src):
        src = src
        dst = os.getcwd()+"/media/file/관리/취합 파일/"
        shutil.copy(src, dst)

    # 파일 리스트를 받아옴
    def ZipDownload(file_list):
        f_path = os.getcwd()+"/media/file/관리"
        new_zips= zipfile.ZipFile(f_path+"/압축 파일/Multizip.zip", 'w')
        for folder, subfolders, files in os.walk(f_path):
            for file in files:
                if file in file_list:
                    new_zips.write(os.path.join(folder, file), os.path.relpath(os.path.join(folder,file), f_path), compress_type = zipfile.ZIP_DEFLATED)
        new_zips.close()
